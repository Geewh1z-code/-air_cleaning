import random
import tkinter as tk
from tkinter import messagebox, ttk
import mysql.connector
from datetime import datetime
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import threading
import statistics
import openpyxl


# Параметры подключения к базе данных
db_config = {
    'host': '127.0.0.1',
    'user': 'your_username',
    'password': 'your_password',
    'database': 'air_quality_monitoring'
}

# диапазоны, устройства, инициализация
ranges = {
    'Temperature': (20, 60),
    'Humidity': (0, 100),
    'PM2.5': (0, 50),
    'PM10': (0, 100),
    'NO2': (0, 200),
    'SO2': (0, 100),
    'CO': (0, 10),
}

target_values = {
    'Temperature': 27,
    'Humidity': 45
}

# Создаем три набора данных для каждого цеха
current_values = {
    1: {key: random.uniform(min_val, max_val) for key, (min_val, max_val) in ranges.items()},  # Цех 1
    2: {key: random.uniform(min_val, max_val) for key, (min_val, max_val) in ranges.items()},  # Цех 2
    3: {key: random.uniform(min_val, max_val) for key, (min_val, max_val) in ranges.items()}   # Цех 3
}

# Состояние спайков для каждого цеха
spike_active = {1: False, 2: False, 3: False}
spike_parameter = {1: None, 2: None, 3: None}
spike_counter = {1: 0, 2: 0, 3: 0}
spike_value = {1: None, 2: None, 3: None}
spike_probability = {1: 0.1, 2: 0.1, 3: 0.1}

# Функции handle_spikes, generate_sensor_data, calculate_aqi, calculate_aqi_percentage
def handle_spikes(current_values, spike_parameter):
    conn = mysql.connector.connect(**db_config)
    cursor = conn.cursor()

    for param in spike_parameter:
        current_value = current_values[param]
        # Находим device_id для параметра
        device_type = None
        for dev, params in {
            'MQ-135': ['NO2', 'SO2', 'CO'],
            'SDS011': ['PM2.5', 'PM10'],
            'DHT22': ['Temperature', 'Humidity']
        }.items():
            if param in params:
                device_type = dev
                break

        # Выбираем случайный location_id (1, 2 или 3)
        location_id = random.randint(1, 3)
        cursor.execute("SELECT device_id FROM devices WHERE device_type = %s AND location_id = %s", (device_type, location_id))
        device_id = cursor.fetchone()[0]

        # Преобразуем параметр в формат, соответствующий ENUM в таблице alerts
        param_mapping = {
            'Temperature': 'temperature',
            'Humidity': 'humidity',
            'PM2.5': 'PM2_5',
            'PM10': 'PM10',
            'NO2': 'NO2',
            'SO2': 'SO2',
            'CO': 'CO',
            'AQI': 'AQI'
        }
        param_enum = param_mapping[param]

        # Вставляем запись в таблицу alerts
        cursor.execute("""
            INSERT INTO alerts (device_id, parameter, value, alert_date)
            VALUES (%s, %s, %s, %s)
        """, (device_id, param_enum, current_value, datetime.now()))

    conn.commit()
    cursor.close()
    conn.close()

# Функция генерации данных (модифицированная)
def generate_sensor_data():
    global current_values, spike_active, spike_parameter, spike_counter, spike_value, spike_probability
    data_to_save = {}

    for location_id in [1, 2, 3]:  # Для каждого цеха
        if not spike_active[location_id]:
            if random.random() < spike_probability[location_id]:
                parameter_choice = random.choices(['Temperature', 'Humidity', 'Pollutants'], weights=[0.33, 0.33, 0.34])[0]
                if parameter_choice == 'Pollutants':
                    pollutants = ['PM2.5', 'PM10', 'NO2', 'SO2', 'CO']
                    spike_parameter[location_id] = random.sample(pollutants, k=random.randint(2, 3))
                else:
                    spike_parameter[location_id] = [parameter_choice]
                spike_value[location_id] = {param: current_values[location_id][param] for param in spike_parameter[location_id]}
                for param in spike_parameter[location_id]:
                    new_value = current_values[location_id][param] * 1.35
                    min_val, max_val = ranges[param]
                    current_values[location_id][param] = min(max(new_value, min_val), max_val)
                spike_active[location_id] = True
                spike_counter[location_id] = 0
                spike_probability[location_id] = min(spike_probability[location_id] + 0.02, 1.0)
                handle_spikes(current_values[location_id], spike_parameter[location_id])
        else:
            spike_counter[location_id] += 1
            if spike_counter[location_id] <= 5:
                for param in spike_parameter[location_id]:
                    difference = current_values[location_id][param] - spike_value[location_id][param]
                    current_values[location_id][param] -= difference / 5
            else:
                spike_active[location_id] = False
                spike_parameter[location_id] = None
                spike_value[location_id] = None
        for key in current_values[location_id]:
            if not spike_active[location_id] or key not in spike_parameter[location_id]:
                change = current_values[location_id][key] * random.uniform(-0.1, 0.1)
                current_values[location_id][key] += change
                if key in target_values:
                    target = target_values[key]
                    difference = target - current_values[location_id][key]
                    current_values[location_id][key] += difference * 0.1
                min_val, max_val = ranges[key]
                current_values[location_id][key] = max(min(current_values[location_id][key], max_val), min_val)
        data_to_save[location_id] = current_values[location_id]
    return data_to_save

def calculate_aqi(data):
    aqi_breakpoints = {
        'PM2.5': [(0, 12, 0, 50), (12, 35.4, 51, 100), (35.4, 55.4, 101, 150), (55.4, 150.4, 151, 200), (150.4, 250.4, 201, 300), (250.4, 500.4, 301, 500)],
        'PM10': [(0, 54, 0, 50), (54, 154, 51, 100), (154, 254, 101, 150), (254, 354, 151, 200), (354, 424, 201, 300), (424, 604, 301, 500)],
        'NO2': [(0, 53, 0, 50), (53, 100, 51, 100), (100, 360, 101, 150), (360, 649, 151, 200), (649, 1249, 201, 300), (1249, 2049, 301, 500)],
        'SO2': [(0, 35, 0, 50), (35, 75, 51, 100), (75, 185, 101, 150), (185, 304, 151, 200), (304, 604, 201, 300), (604, 1004, 301, 500)],
        'CO': [(0, 4.4, 0, 50), (4.4, 9.4, 51, 100), (9.4, 12.4, 101, 150), (12.4, 15.4, 151, 200), (15.4, 30.4, 201, 300), (30.4, 50.4, 301, 500)]
    }
    aqi_values = []
    for param, value in data.items():
        if param in aqi_breakpoints:
            for (c_low, c_high, i_low, i_high) in aqi_breakpoints[param]:
                if c_low <= value <= c_high:
                    aqi = ((i_high - i_low) / (c_high - c_low)) * (value - c_low) + i_low
                    aqi_values.append(aqi)
                    break
    return max(aqi_values) if aqi_values else 0

def calculate_aqi_percentage(aqi):
    return max(0, 100 - (aqi / 500) * 100)

# Функция для получения последних 20 записей из базы данных
def get_last_20_records(location_id):
    conn = mysql.connector.connect(**db_config)
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT timestamp, temperature, humidity, PM2_5, PM10, NO2, SO2, CO, AQI, air_quality
        FROM air_quality_data
        WHERE location_id = %s
        ORDER BY timestamp DESC
        LIMIT 20
    """, (location_id,))
    records = cursor.fetchall()
    cursor.close()
    conn.close()
    return records


def get_last_100_records(location_id):
    conn = mysql.connector.connect(**db_config)
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT timestamp, temperature, humidity, PM2_5, PM10, NO2, SO2, CO, AQI
        FROM air_quality_data
        WHERE location_id = %s
        ORDER BY timestamp DESC
        LIMIT 100
    """, (location_id,))
    records = cursor.fetchall()
    cursor.close()
    conn.close()
    return records

# Функция для логирования действий
def log_action(user_id, action):
    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO logs (user_id, action, timestamp)
            VALUES (%s, %s, %s)
        """, (user_id, action, datetime.now()))
        conn.commit()
        cursor.close()
        conn.close()
    except mysql.connector.Error as err:
        print(f"Ошибка при логировании: {err}")


# Функция для логирования действий
def log_action(user_id, action):
    try:
        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor()
        cursor.execute("""
            INSERT INTO logs (user_id, action, timestamp)
            VALUES (%s, %s, %s)
        """, (user_id, action, datetime.now()))
        conn.commit()
        cursor.close()
        conn.close()
    except mysql.connector.Error as err:
        print(f"Ошибка при логировании: {err}")

# Класс окна авторизации
class LoginWindow:
    def __init__(self, root):
        self.root = root
        self.root.title("Авторизация")
        self.root.geometry("300x200")
        self.root.resizable(False, False)

        self.frame = tk.Frame(self.root)
        self.frame.pack(pady=20)

        self.label_username = tk.Label(self.frame, text="Логин:", font=("Arial", 12))
        self.label_username.pack()
        self.entry_username = tk.Entry(self.frame, font=("Arial", 12))
        self.entry_username.pack(pady=5)

        self.label_password = tk.Label(self.frame, text="Пароль:", font=("Arial", 12))
        self.label_password.pack()
        self.entry_password = tk.Entry(self.frame, font=("Arial", 12), show="*")
        self.entry_password.pack(pady=5)

        self.login_button = tk.Button(self.frame, text="Войти", font=("Arial", 12), command=self.check_login)
        self.login_button.pack(pady=10)

        self.root.bind("<Return>", lambda event: self.check_login())

    def check_login(self):
        username = self.entry_username.get()
        password = self.entry_password.get()

        if not username or not password:
            messagebox.showerror("Ошибка", "Пожалуйста, введите логин и пароль")
            return

        try:
            conn = mysql.connector.connect(**db_config)
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT user_id, username FROM users WHERE username = %s AND password_hash = %s", (username, password))
            user = cursor.fetchone()

            if user:
                user_id = user['user_id']
                # Логируем вход
                log_action(user_id, 'login')
                self.root.destroy()
                self.open_main_app(user['username'], user_id)
            else:
                messagebox.showerror("Ошибка", "Неверный логин или пароль")

            cursor.close()
            conn.close()

        except mysql.connector.Error as err:
            messagebox.showerror("Ошибка", f"Ошибка подключения к базе данных: {err}")

    def open_main_app(self, username, user_id):
        root = tk.Tk()
        app = SensorApp(root, username=username, user_id=user_id)
        root.mainloop()



# Основное приложение
class SensorApp:
    def __init__(self, root, username="admin", user_id=None):
        self.root = root
        self.username = username
        self.user_id = user_id
        self.root.title("Мониторинг параметров завода")
        self.root.geometry("1400x800")

        self.main_frame = tk.Frame(self.root)
        self.main_frame.pack(fill=tk.BOTH, expand=True)

        self.left_frame = tk.Frame(self.main_frame)
        self.left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.bottom_left_frame = tk.Frame(self.left_frame)
        self.bottom_left_frame.pack(side=tk.LEFT, fill=tk.Y)

        self.fig_left = plt.Figure(figsize=(4, 6))
        self.axes_left = [self.fig_left.add_subplot(4, 1, i+1) for i in range(4)]
        self.fig_left.subplots_adjust(left=0.1, right=0.95, top=0.95, bottom=0.05, hspace=0.5)

        self.canvas_left = FigureCanvasTkAgg(self.fig_left, master=self.bottom_left_frame)
        self.canvas_left.get_tk_widget().pack(fill=tk.BOTH, expand=True)

        self.top_left_frame = tk.Frame(self.left_frame)
        self.top_left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10)

        self.user_label = tk.Label(self.top_left_frame, text=f"Пользователь: {self.username}", font=("Arial", 14))
        self.user_label.pack(pady=10)

        self.location_label = tk.Label(self.top_left_frame, text="Выберите цех:", font=("Arial", 14))
        self.location_label.pack(pady=10)

        self.location_buttons_frame = tk.Frame(self.top_left_frame)
        self.location_buttons_frame.pack(pady=10)

        style = ttk.Style()
        style.configure("Active.TButton", background="lightgreen", foreground="black")
        style.configure("Inactive.TButton", background="lightgrey", foreground="black")

        self.location_buttons = {}
        for location in ["Цех 1", "Цех 2", "Цех 3"]:
            btn = ttk.Button(
                self.location_buttons_frame,
                text=location,
                style="Inactive.TButton",
                command=lambda loc=location: self.update_location(loc)
            )
            btn.pack(side=tk.LEFT, padx=5)
            self.location_buttons[location] = btn

        self.location_buttons["Цех 1"].configure(style="Active.TButton")

        self.time_label = tk.Label(self.top_left_frame, text="Текущее время: --:--:--", font=("Arial", 14))
        self.time_label.pack(pady=10)

        self.air_quality_label = tk.Label(self.top_left_frame, text="Качество воздуха: --%", font=("Arial", 14))
        self.air_quality_label.pack(pady=10)

        self.button_frame = tk.Frame(self.top_left_frame)
        self.button_frame.pack(pady=10)

        self.reports_button = tk.Button(self.button_frame, text="Просмотреть отчеты", command=self.view_reports, font=("Arial", 12))
        self.reports_button.pack(side=tk.LEFT, padx=5)

        self.report_button = tk.Button(self.button_frame, text="Создать отчет", command=self.create_report, font=("Arial", 12))
        self.report_button.pack(side=tk.LEFT, padx=5)

        self.exit_button = tk.Button(self.button_frame, text="Выход", command=self.on_closing, font=("Arial", 12))
        self.exit_button.pack(side=tk.LEFT, padx=5)

        self.right_frame = tk.Frame(self.main_frame)
        self.right_frame.pack(side=tk.RIGHT, fill=tk.Y)

        self.fig_right = plt.Figure(figsize=(4, 6))
        self.axes_right = [self.fig_right.add_subplot(4, 1, i+1) for i in range(4)]
        self.fig_right.subplots_adjust(left=0.1, right=0.95, top=0.95, bottom=0.05, hspace=0.5)

        self.canvas_right = FigureCanvasTkAgg(self.fig_right, master=self.right_frame)
        self.canvas_right.get_tk_widget().pack(fill=tk.BOTH, expand=True)

        self.parameters_left = ['NO2', 'SO2', 'CO', 'AQI']
        self.parameters_right = ['temperature', 'humidity', 'PM2_5', 'PM10']
        self.all_parameters = self.parameters_left + self.parameters_right

        self.current_location_id = 1
        self.is_updating = False
        self.running = True
        self.update_data()
        self.update_time()

    def create_report(self):
        # Логируем генерацию отчета
        if self.user_id:
            log_action(self.user_id, 'report_generated')

            # Добавляем запись в таблицу reports
            try:
                conn = mysql.connector.connect(**db_config)
                cursor = conn.cursor()
                if self.user_id:
                    cursor.execute("""
                            INSERT INTO reports (report_date, user_id)
                            VALUES (%s, %s)
                        """, (datetime.now(), self.user_id))
                conn.commit()
                cursor.close()
                conn.close()
            except mysql.connector.Error as err:
                messagebox.showerror("Ошибка", f"Не удалось сохранить отчет в базе данных: {err}")
                return

        report_window = tk.Toplevel(self.root)
        report_window.title("Отчет по параметрам")
        report_window.geometry("800x600")

        tree = ttk.Treeview(report_window, columns=("Parameter", "Average", "Max", "Min"), show="headings")
        tree.heading("Parameter", text="Параметр")
        tree.heading("Average", text="Среднее")
        tree.heading("Max", text="Максимум")
        tree.heading("Min", text="Минимум")

        tree.column("Parameter", width=150)
        tree.column("Average", width=100)
        tree.column("Max", width=100)
        tree.column("Min", width=100)

        scrollbar = ttk.Scrollbar(report_window, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscroll=scrollbar.set)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        for location_id in [1, 2, 3]:
            records = get_last_100_records(location_id)
            if not records:
                tree.insert("", tk.END, values=(f"Цех {location_id} (нет данных)", "", "", ""))
                continue

            tree.insert("", tk.END, values=(f"Цех {location_id}", "", "", ""))

            data = {
                'temperature': [record['temperature'] for record in records],
                'humidity': [record['humidity'] for record in records],
                'PM2_5': [record['PM2_5'] for record in records],
                'PM10': [record['PM10'] for record in records],
                'NO2': [record['NO2'] for record in records],
                'SO2': [record['SO2'] for record in records],
                'CO': [record['CO'] for record in records],
                'AQI': [record['AQI'] for record in records]
            }

            for param in self.all_parameters:
                values = data[param]
                if values:
                    avg = round(statistics.mean(values), 2)
                    max_val = max(values)
                    min_val = min(values)
                    tree.insert("", tk.END, values=(param, avg, max_val, min_val))
                else:
                    tree.insert("", tk.END, values=(param, "Нет данных", "Нет данных", "Нет данных"))

            tree.insert("", tk.END, values=("", "", "", ""))

        save_button = tk.Button(report_window, text="Скачать в Excel", command=lambda: self.save_to_excel(tree))
        save_button.pack(pady=10)

    def save_to_excel(self, tree):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Отчет по параметрам"

        headers = ["Параметр", "Среднее", "Максимум", "Минимум"]
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num).value = header

        row_num = 2
        for item in tree.get_children():
            values = tree.item(item)["values"]
            for col_num, value in enumerate(values, 1):
                sheet.cell(row=row_num, column=col_num).value = value
            row_num += 1

        try:
            workbook.save("report.xlsx")
            messagebox.showinfo("Успех", "Отчет сохранен в report.xlsx")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось сохранить файл: {e}")

    def update_location(self, value):
        location_mapping = {
            "Цех 1": 1,
            "Цех 2": 2,
            "Цех 3": 3
        }
        self.current_location_id = location_mapping[value]

        for location, btn in self.location_buttons.items():
            if location == value:
                btn.configure(style="Active.TButton")
            else:
                btn.configure(style="Inactive.TButton")

        print(f"Выбран цех: {value}, ID: {self.current_location_id}")
        self.plot_data()

    def update_time(self):
        if not self.running:
            return
        current_time = datetime.now().strftime('%H:%M:%S')
        self.time_label.config(text=f"Текущее время: {current_time}")
        self.root.after(1000, self.update_time)

    def plot_data(self):
        records = get_last_20_records(self.current_location_id)
        if not records:
            for ax in self.axes_left:
                ax.clear()
                ax.set_title(ax.get_title(), fontsize=10)
                ax.set_xlabel('Время', fontsize=8)
                ax.set_ylabel('')
                ax.grid(True)
                ax.tick_params(axis='x', labelbottom=False)
            for ax in self.axes_right:
                ax.clear()
                display_title = ax.get_ylabel()
                ax.set_xlabel('Время', fontsize=8)
                ax.yaxis.set_label_position("left")
                ax.set_ylabel(display_title, fontsize=10, rotation=0, labelpad=40)
                ax.grid(True)
                ax.tick_params(axis='x', labelbottom=False)
            self.air_quality_label.config(text="Качество воздуха: --%")
            self.canvas_left.draw()
            self.canvas_right.draw()
            return

        timestamps = [record['timestamp'].strftime('%H:%M:%S') for record in records]
        timestamps = timestamps[::-1]
        data = {
            'temperature': [record['temperature'] for record in records][::-1],
            'humidity': [record['humidity'] for record in records][::-1],
            'PM2_5': [record['PM2_5'] for record in records][::-1],
            'PM10': [record['PM10'] for record in records][::-1],
            'NO2': [record['NO2'] for record in records][::-1],
            'SO2': [record['SO2'] for record in records][::-1],
            'CO': [record['CO'] for record in records][::-1],
            'AQI': [record['AQI'] for record in records][::-1],
            'air_quality': [record['air_quality'] for record in records][::-1]
        }

        latest_air_quality = data['air_quality'][-1] if data['air_quality'] else "--"
        self.air_quality_label.config(text=f"Качество воздуха: {latest_air_quality}%")

        # Проверка на превышение пороговых значений (для alert_triggered)
        thresholds = {
            'NO2': 100,
            'SO2': 50,
            'CO': 10,
            'AQI': 150,
            'temperature': 35,
            'humidity': 90,
            'PM2_5': 50,
            'PM10': 100
        }

        for param in self.all_parameters:
            latest_value = data[param][-1] if data[param] else None
            if latest_value and latest_value > thresholds[param]:
                log_action(self.user_id, 'alert_triggered')
                print(f"Алерт: {param} превысил порог ({latest_value} > {thresholds[param]})")

        for i, ax in enumerate(self.axes_left):
            ax.clear()
            param = self.parameters_left[i]
            color = 'red' if spike_active[self.current_location_id] and param in spike_parameter[self.current_location_id] else 'blue'
            ax.plot(timestamps, data[param], marker='o', linestyle='-', color=color, markersize=5, linewidth=2)
            ax.set_xlabel(param, fontsize=10)
            ax.tick_params(axis='x', rotation=45, labelsize=6, labelbottom=False)
            ax.tick_params(axis='y', labelsize=6)
            ax.grid(True)

        self.canvas_left.draw()
        self.canvas_left.flush_events()

        for i, ax in enumerate(self.axes_right):
            ax.clear()
            param = self.parameters_right[i]
            color = 'red' if spike_active[self.current_location_id] and param in spike_parameter[self.current_location_id] else 'blue'
            ax.plot(timestamps, data[param], marker='o', linestyle='-', color=color, markersize=5, linewidth=2)
            ax.set_xlabel(param, fontsize=10)
            ax.tick_params(axis='x', rotation=45, labelsize=6, labelbottom=False)
            ax.tick_params(axis='y', labelsize=6)
            ax.grid(True)

        self.canvas_right.draw()
        self.canvas_right.flush_events()

    def update_data(self):
        if not self.running or self.is_updating:
            return

        self.is_updating = True

        def background_task():
            try:
                data_to_save = generate_sensor_data()
                conn = mysql.connector.connect(**db_config)
                cursor = conn.cursor()

                for location_id, data in data_to_save.items():
                    aqi = calculate_aqi(data)
                    aqi_percentage = calculate_aqi_percentage(aqi)
                    cursor.execute("SELECT device_id FROM devices WHERE location_id = %s LIMIT 1", (location_id,))
                    device_id = cursor.fetchone()[0]
                    cursor.execute("""
                        INSERT INTO air_quality_data (device_id, location_id, timestamp, temperature, humidity, PM2_5, PM10, NO2, SO2, CO, AQI, air_quality)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (device_id, location_id, datetime.now(), data['Temperature'], data['Humidity'], data['PM2.5'], data['PM10'], data['NO2'], data['SO2'], data['CO'], aqi, aqi_percentage))

                conn.commit()
                cursor.close()
                conn.close()

                self.root.after(0, self.plot_data)

            finally:
                self.is_updating = False

        threading.Thread(target=background_task, daemon=True).start()
        self.root.after(3000, self.update_data)

    def on_closing(self):
        # Логируем выход
        if self.user_id:
            log_action(self.user_id, 'logout')
        self.running = False
        self.root.destroy()

    def view_reports(self):
        reports_window = tk.Toplevel(self.root)
        reports_window.title("Список отчетов")
        reports_window.geometry("600x400")

        tree = ttk.Treeview(reports_window, columns=("Report ID", "Date", "User ID"), show="headings")
        tree.heading("Report ID", text="ID отчета")
        tree.heading("Date", text="Дата")
        tree.heading("User ID", text="ID пользователя")

        tree.pack(fill=tk.BOTH, expand=True)

        conn = mysql.connector.connect(**db_config)
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT report_id, report_date, user_id FROM reports WHERE user_id = %s", (self.user_id,))
        reports = cursor.fetchall()
        for report in reports:
            tree.insert("", tk.END, values=(report['report_id'], report['report_date'], report['user_id']))
        cursor.close()
        conn.close()


# Запуск приложения
if __name__ == "__main__":
    root = tk.Tk()
    login_app = LoginWindow(root)
    root.mainloop()
